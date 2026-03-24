import statistics
import urllib
import openpyxl
import requests
import urllib.parse

# function for FMP

API_KEY = "un9SnAnMNfh3V2e1C1MSdntTOqefUvoF"


def get_endpoint(endpoint, parameters):
    url = f"https://financialmodelingprep.com/stable/{endpoint}"
    params = parameters.copy()
    params["apikey"] = API_KEY

    print("Getting Endpoint:", url + "?" + urllib.parse.urlencode(params))
    response = requests.get(url, params=params)
    response.raise_for_status()
    return response.json()

# create workbook for companies
stock_wb = openpyxl.Workbook()
ws1 = stock_wb.active
ws1.title = "Summary"  # Renaming default sheet

# add worksheets
company_ws = stock_wb.create_sheet("Company Data")
stock_ws = stock_wb.create_sheet("Stock Data")

# add headers
company_ws.append(["Symbol", "Company Name", "Sector", "Exchange", "Volatility", "Trend Slope"])
stock_ws.append(["Symbol", "Date", "Open", "Close"])

# read in stocklist file
ticker_list = []
try:
    with open("stocklist.txt") as file_pointer:
        file_contents = file_pointer.read()

    for line_value in file_contents.splitlines():
        # Adjust logic if your file format differs
        if "CD" in line_value:
            ticker_list = line_value[3:].split(",")
            break
except FileNotFoundError:
    print("Error: stocklist.txt not found. Using default list.")
    ticker_list = ["CCL", "AAPL", "MSFT"]

print("Tickers to process:", ticker_list)

# Loop through tickers
for ticker in ticker_list:
    ticker = ticker.strip() # This removes ' NVDA' -> 'NVDA'
    if not ticker: continue # Skips empty strings
    print(f"\n--- Processing: {ticker} ---")

    # 1. Download profile data
    profile_request = get_endpoint("profile", {"symbol": ticker})

    if isinstance(profile_request, list) and len(profile_request) > 0:
        company_profile = profile_request[0]
        company_name = company_profile.get("companyName", "N/A")
        sector = company_profile.get("sector", "N/A")
        exchange = company_profile.get("exchange", "N/A")
    else:
        print(f"No profile data found for {ticker}")
        continue

    # 2. Download stock price data
    price_request = get_endpoint("historical-price-eod/full", {"symbol": ticker})

    if isinstance(price_request, list):
        historical_list = price_request
    elif isinstance(price_request, dict):
        historical_list = price_request.get("historical", [])
    else:
        print(f"Skipping {ticker}: unexpected API response.")
        continue

    if not historical_list:
        print(f"No historical price data found for {ticker}")
        continue

    # Setup for math/excel
    close_price_list = []
    x_values = []

    # Loop through the last 30 days
    for index, day in enumerate(historical_list[:30]):
        date = day.get("date")
        open_p = day.get("open")
        close_p = day.get("close")

        # Append to stock worksheet
        stock_ws.append([ticker, date, open_p, close_p])

        # Collect values for calculations
        close_price_list.append(close_p)
        x_values.append(index)

    # 3. Calculations
    # Volatility (Standard Deviation)
    if len(close_price_list) > 1:
        volatility = statistics.stdev(close_price_list)
        # Slope: (Latest Price - Oldest Price) / (Latest Day - Oldest Day)
        # Note: Index 0 is usually the most recent day in FMP
        slope = (close_price_list[0] - close_price_list[-1]) / (x_values[-1] - x_values[0])
    else:
        volatility = 0
        slope = 0

    # 4. Append to company worksheet
    company_ws.append([
        ticker,
        company_name,
        sector,
        exchange,
        round(volatility, 2),
        round(slope, 2)
    ])

# save workbook
stock_wb.save("project3.xlsx")
print("\nSuccess! Data saved to project3.xlsx")
